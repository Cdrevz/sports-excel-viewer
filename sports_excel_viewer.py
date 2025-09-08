import streamlit as st
import pandas as pd
from io import BytesIO
import polars as pl
import re
from datetime import datetime
import xlrd
from openpyxl import Workbook
import tempfile
import os

st.set_page_config(
    page_title="Sports Excel Viewer",
    page_icon="🏆",
    layout="wide"
)

st.sidebar.title("Navigation")
page = st.sidebar.radio("Select", [
    "Ice Hockey",
    "Soccer",
    "Rugby",
    "Basketball",
    "Aussie Rules",
    "Program Review"
    ])

def convert_xls_to_xlsx(uploaded_file):
    """Convert .xls file to .xlsx format using xlrd and openpyxl"""
    try:
        # Save uploaded file to temporary location
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            tmp_xls_path = tmp_file.name
        
        # Read the .xls file using xlrd
        book = xlrd.open_workbook(tmp_xls_path)
        sheet = book.sheet_by_index(0)
        
        # Create a new .xlsx file using openpyxl
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        
        # Write data to the new .xlsx file
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                ws.cell(row=row+1, column=col+1, value=sheet.cell_value(row, col))
        
        # Save to temporary .xlsx file
        tmp_xlsx_path = tmp_xls_path.replace('.xls', '.xlsx')
        wb.save(tmp_xlsx_path)
        
        # Clean up the temporary .xls file
        os.unlink(tmp_xls_path)
        
        # Return the path to the converted file
        return tmp_xlsx_path
        
    except Exception as e:
        st.error(f"Error converting .xls file: {str(e)}")
        return None

def process_excel(uploaded_file):
    st.success("Excel file uploaded successfully!")
    st.write("File details:")
    file_details = {"Filename": uploaded_file.name, "FileType": uploaded_file.type, "FileSize": uploaded_file.size}
    st.json(file_details)

if page == "Ice Hockey":
    st.title("🏒 Ice Hockey Excel Upload")
    uploaded_file = st.file_uploader("Upload Excel file for Ice Hockey", type=["xls", "xlsx"])
    if uploaded_file is not None:
        # Convert .xls to .xlsx if needed
        if uploaded_file.name.endswith('.xls'):
            st.info("Converting .xls file to .xlsx format...")
            converted_file_path = convert_xls_to_xlsx(uploaded_file)
            if converted_file_path is None:
                st.error("Failed to convert .xls file. Please try again.")
                st.stop()
            # Read the converted file
            df = pl.read_excel(converted_file_path)
            # Clean up the temporary converted file
            try:
                os.unlink(converted_file_path)
            except:
                pass
        else:
            df = pl.read_excel(uploaded_file)
        new_columns = df.head(1).row(0)
        df = df.slice(0)
        df.columns = new_columns
        df = df.with_columns(
            pl.when(
                df["Date"].str.starts_with("Ice Hockey")
            )
            .then(df["Date"])
            .otherwise(None)
            .alias("League")
        ).with_columns(
            pl.col("League").forward_fill()
        )
        
        # First filter out unwanted leagues (women's leagues and Liiga Relegation)
        if "League" in df.columns:
            df = df.filter(
                ~(
                    pl.col("League").str.to_lowercase().str.contains("liiga, relegation/promotion") 
                )
            )
        
        df = df.filter(pl.col("Postponed") == "0")
        
        # Then filter for wanted leagues
        filter_words = [
            "Russia.KHL", 
            "Czechia.Extraliga", 
            "Slovakia.Extraliga", 
            "Sweden.SHL",
            "Finland.Liiga",
            "Champions Hockey League", 
            "International.U20 World Championship, Group",
            "International.World Championship, Group",
            "International.World Championship, Knockout Stage"
        ]
        df = df.filter(
            pl.col("League").str.contains("|".join(filter_words))
        )
        
        # Clean up league names
        df = df.with_columns(
            pl.col("League")
            .str.replace(r",", "")
            .str.replace(r"(?i)\bweek\b", "")
            .str.replace("Ice Hockey.", "")
            .str.replace("Playoff,", "")
            .str.replace("Playoffs,", "")
            .str.replace("Playout", "")
            .str.replace("Knockout Stage,", "")
            .str.strip_chars()
            .map_elements(
                lambda x: re.sub(r'\d+', '', str(x)) if x is not None else None,
                return_dtype=pl.Utf8
            )
            .alias("League")
        )
        
        # Rest of your processing code remains the same...
        df = df.with_columns(
            pl.when(pl.col("AP").is_not_null())
            .then(
                pl.col("AP")
                .str.split(":")
                .map_elements(lambda x: sum(int(i) for i in x), return_dtype=pl.Int64)
            )
            .when(pl.col("OT").is_not_null())
            .then(
                pl.col("OT")
                .str.split(":")
                .map_elements(lambda x: sum(int(i) for i in x), return_dtype=pl.Int64)
            )
            .when(pl.col("FT").is_not_null())
            .then(
                pl.col("FT")
                .str.split(":")
                .map_elements(lambda x: sum(int(i) for i in x), return_dtype=pl.Int64)
            )
            .otherwise(None)
            .alias("Goals")
        )

        df = df.with_columns(
            pl.when(pl.col("AP").is_not_null())
            .then(5)
            .when(pl.col("OT").is_not_null())
            .then(4)
            .otherwise(3)
            .alias("Period")
        )

        columns_to_drop = ["FT", "1", "2", "3", "OT", "AP", "Postponed"]
        df = df.drop(columns_to_drop)

        df = df.with_columns(
            pl.lit(None).alias("Datapoints"),
            pl.lit(None).alias("Issue"), 
            pl.lit(None).alias("Suspensions"),
            pl.lit(None).alias("Suspension issue"),  
            pl.lit(None).alias("Goals issue"),    
        )
        
        if "Goals" in df.columns:
            df = df.filter(pl.col("Goals").is_not_null())
            
        df_display = df.select(["Date", "KO", "League", "Home", "Away", "Match Id", "Datapoints", "Issue", "Goals","Goals issue","Suspensions","Suspension issue","Period"])
        
        st.subheader("Processed Ice Hockey Data")
        st.dataframe(df_display)
        current_date = datetime.now().strftime("%Y%m%d")
        output = BytesIO()
        df_display.write_excel(output)
        output.seek(0)
        st.download_button(
            label="Download Excel",
            data=output,
            file_name=f"Ice Hockey - {current_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

elif page == "Soccer":
    st.title("⚽ Soccer Excel Upload")
    uploaded_file = st.file_uploader("Upload Excel file for Soccer", type=["xls", "xlsx"])
    if uploaded_file is not None:
        try:
            # Convert .xls to .xlsx if needed
            if uploaded_file.name.endswith('.xls'):
                st.info("Converting .xls file to .xlsx format...")
                converted_file_path = convert_xls_to_xlsx(uploaded_file)
                if converted_file_path is None:
                    st.error("Failed to convert .xls file. Please try again.")
                    st.stop()
                # Read the converted file
                df = pl.read_excel(converted_file_path)
                # Clean up the temporary converted file
                try:
                    os.unlink(converted_file_path)
                except:
                    pass
            else:
                df = pl.read_excel(uploaded_file)
            new_columns = df.head(1).row(0)
            df = df.slice(0)
            df.columns = new_columns
            df = df.with_columns(
                pl.when(
                    df["Date"].str.starts_with("Soccer")
                )
                .then(df["Date"])
                .otherwise(None)
                .alias("League")
            ).with_columns(
                pl.col("League").forward_fill()
            )
            if "League" in df.columns:
                df = df.filter(
                    ~(
                        pl.col("League").str.to_lowercase().str.contains("women") |
                        pl.col("League").str.to_lowercase().str.contains("Spain.LaLiga 2") |
                        pl.col("League").str.contains("MLS Next Pro")
                    )
                )
            df = df.filter(pl.col("Postponed") == "0")
            filter_words = ["Italy.Serie A", "Spain.LaLiga", "England.Premier League", "Germany.Bundesliga", "USA.Major League Soccer","Austria.Bundesliga","USA.MLS","International Clubs.UEFA Champions League"]
            df = df.filter(
                pl.col("League").str.contains("|".join(filter_words))
            )
            df = df.with_columns(
                pl.col("League")
                .str.replace(r",", "")
                .str.replace(r"(?i)\bweek\b", "")
                .str.replace("Soccer.", "")
                .str.strip_chars()
                .map_elements(
                    lambda x: re.sub(r'\d+', '', str(x)) if x is not None else None,
                    return_dtype=pl.Utf8
                )
                .alias("League")
            )
            
            columns_to_drop = ["AP","OT","HT","FT","Comment", "Postponed"]
            df = df.drop(columns_to_drop)
            if "Date" in df.columns:
                df = df.with_columns(
                    pl.col("Date").str.strptime(pl.Date, format="%d/%m %y")
                            .dt.strftime("%m/%d/%Y")
                            .alias("Date")
                )
            st.subheader("Processed League Data")
            df_display = df.select([
                "Date", "KO", "League", "Home", "Away", "Match Id"
            ])
            
            st.dataframe(df_display)
            current_date = datetime.now().strftime("%Y%m%d")
            
            # Download button - use df_display instead of df
            output = BytesIO()
            df_display.write_excel(output)  # Write the displayed columns only
            output.seek(0)
            st.download_button(
                label="Download Excel",
                data=output,
                file_name=f"League Data - {current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
                

elif page == "Rugby":
    st.title("🏉 Rugby Excel Upload")
    uploaded_file = st.file_uploader("Upload Excel file for Rugby", type=["xls", "xlsx"])
    if uploaded_file is not None:
        try:
            # Convert .xls to .xlsx if needed
            if uploaded_file.name.endswith('.xls'):
                st.info("Converting .xls file to .xlsx format...")
                converted_file_path = convert_xls_to_xlsx(uploaded_file)
                if converted_file_path is None:
                    st.error("Failed to convert .xls file. Please try again.")
                    st.stop()
                # Read the converted file
                df = pl.read_excel(converted_file_path)
                # Clean up the temporary converted file
                try:
                    os.unlink(converted_file_path)
                except:
                    pass
            else:
                df = pl.read_excel(uploaded_file)
            new_columns = df.head(1).row(0)
            df = df.slice(0)
            df.columns = new_columns
            df = df.with_columns(
                pl.when(
                    df["Date"].str.starts_with("Rugby")
                )
                .then(df["Date"])
                .otherwise(None)
                .alias("League")
            ).with_columns(
                pl.col("League").forward_fill()
            )
            df = df.filter(pl.col("Postponed") == "0")
            filter_words = [
                  "Six Nations",
                  "Super Rugby",
                    "Premiership Rugby",
                    "European Rugby Champions Cup",
                    "The Rugby Championship"
                    ]
            df = df.filter(
                pl.col("League").str.contains("|".join(filter_words))
            )
            df = df.with_columns(
                pl.col("League")
                .str.replace(r",", "")
                .str.replace(r"(?i)\bweek\b", "")
                .str.replace("Rugby.", "")
                .str.strip_chars()
                .map_elements(
                    lambda x: re.sub(r'\d+', '', str(x)) if x is not None else None,
                    return_dtype=pl.Utf8
                )
                .alias("League")
            )
            # Filter out rows where League column contains "women" (case insensitive)
            if "League" in df.columns:
                if "League" in df.columns:
                    df = df.filter(
                        ~(
                            pl.col("League").str.to_lowercase().str.contains("women") |
                            pl.col("League").str.contains("Premiership Rugby Cup Playoffs") |
                            pl.col("League").str.contains("U Six Nations") |
                            pl.col("League").str.contains("Super Rugby Americas") 

                        )
                    )
            columns_to_drop = ["AP","OT","HT","FT","Comment", "Postponed"]
            df = df.drop(columns_to_drop)
            if "Date" in df.columns:
                df = df.with_columns(
                    pl.col("Date").str.strptime(pl.Date, format="%d/%m %y")
                            .dt.strftime("%m/%d/%Y")
                            .alias("Date")
                )
            st.subheader("Processed Rugby Data")
            df_display = df.select([
                "Date", "KO", "League", "Home", "Away", "Match Id"
            ])
            
            st.dataframe(df_display)
            current_date = datetime.now().strftime("%Y%m%d")
            
            # Download button - use df_display instead of df
            output = BytesIO()
            df_display.write_excel(output)  # Write the displayed columns only
            output.seek(0)
            st.download_button(
                label="Download Excel",
                data=output,
                file_name=f"Rugby - {current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")

elif page == "Basketball":
    st.title("🏀 Basketball Excel Upload")
    uploaded_file = st.file_uploader("Upload Excel file for Basketball", type=["xls", "xlsx"])
    if uploaded_file is not None:
        try:
            # Convert .xls to .xlsx if needed
            if uploaded_file.name.endswith('.xls'):
                st.info("Converting .xls file to .xlsx format...")
                converted_file_path = convert_xls_to_xlsx(uploaded_file)
                if converted_file_path is None:
                    st.error("Failed to convert .xls file. Please try again.")
                    st.stop()
                # Read the converted file
                df = pl.read_excel(converted_file_path)
                # Clean up the temporary converted file
                try:
                    os.unlink(converted_file_path)
                except:
                    pass
            else:
                df = pl.read_excel(uploaded_file)
            new_columns = df.head(1).row(0)
            df = df.slice(0)
            df.columns = new_columns
            df = df.with_columns(
                pl.when(
                    df["Date"].str.starts_with("Basketball")
                )
                .then(df["Date"])
                .otherwise(None)
                .alias("League")
            ).with_columns(
                pl.col("League").forward_fill()
            )
            df = df.filter(pl.col("Postponed") == "0")
            filter_words = ["Italy.Serie A",
                            "France.LNB Elite",
                            "Turkiye.Super Lig",
                            "Spain.Liga ACB",
                            "Germany.BBL",
                            "International.Euroleague",
                            "International.Eurocup",
                            "Israel.Super League",
                            "International.ABA Liga",
                            "China.CBA",
                            "Australia.NBL",
                            "Greece.Greek Basketball League",
                            "International.FIBA World Cup",
                            "International.Champions League",
                            "International.ABA Liga",
                            "International.Olympic",
                            "European Championship"
                            ]
            df = df.filter(
                pl.col("League").str.contains("|".join(filter_words))
            )
            df = df.with_columns(
                pl.col("League")
                .str.replace(r",", "")
                .str.replace(r"Playoffs,", "Playoffs")
                .str.replace(r"(?i)\bweek\b", "")
                .str.replace("Basketball.", "")
                .str.strip_chars()
                .map_elements(
                    lambda x: re.sub(r'\d+', '', str(x)) if x is not None else None,
                    return_dtype=pl.Utf8
                )
                .alias("League")
            )
            # Filter out rows where League column contains "women" (case insensitive)
            if "League" in df.columns:
                if "League" in df.columns:
                    df = df.filter(
                        ~(
                            pl.col("League").str.to_lowercase().str.contains("women") |
                            pl.col("League").str.to_lowercase().str.contains("Promotion") |
                            pl.col("League").str.contains("NBL Central") |
                            pl.col("League").str.contains("NBL East") |
                            pl.col("League").str.contains("NBL West") |
                            pl.col("League").str.contains("NBL North") |
                            pl.col("League").str.contains("NBL South") |
                            pl.col("League").str.contains("Champions League Asia Group C") |
                            pl.col("League").str.contains("Champions League Asia Group") |
                            pl.col("League").str.contains("Champions League Asia Group A") |
                            pl.col("League").str.contains("Champions League Asia Group B") |
                            pl.col("League").str.contains("Champions League Asia Group D") |
                            pl.col("League").str.contains("Champions League Asia Group E") |
                            pl.col("League").str.contains("Champions League Asia Group F") |
                            pl.col("League").str.contains("Champions League Asia Group G")  |
                            pl.col("League").str.contains("Champions League Asia Knockout Stage,") |
                            pl.col("League").str.contains("ABA Liga Relegation/Promotion Playoff,") |
                            pl.col("League").str.contains("FIBA World Cup Americas Pre-Qualifiers,")


                        )
                    )
            columns_to_drop = ["1","2","3","4","OT","FT","Comment", "Postponed"]
            df = df.drop(columns_to_drop)
            if "Date" in df.columns:
                df = df.with_columns(
                    pl.col("Date").str.strptime(pl.Date, format="%d/%m %y")
                            .dt.strftime("%m/%d/%Y")
                            .alias("Date")
                )
            st.subheader("Processed Basketball Data")
            df_display = df.select([
                "Date", "KO", "League", "Home", "Away", "Match Id"
            ])
            
            st.dataframe(df_display)
            current_date = datetime.now().strftime("%Y%m%d")
            
            # Download button - use df_display instead of df
            output = BytesIO()
            df_display.write_excel(output)  # Write the displayed columns only
            output.seek(0)
            st.download_button(
                label="Download Excel",
                data=output,
                file_name=f"Basketball - {current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            
elif page == "Aussie Rules":
    st.title("🏈 Aussie Rules Excel Upload")
    uploaded_file = st.file_uploader("Upload Excel file for Aussie Rules", type=["xls", "xlsx"])
    if uploaded_file is not None:
        try:
            # Convert .xls to .xlsx if needed
            if uploaded_file.name.endswith('.xls'):
                st.info("Converting .xls file to .xlsx format...")
                converted_file_path = convert_xls_to_xlsx(uploaded_file)
                if converted_file_path is None:
                    st.error("Failed to convert .xls file. Please try again.")
                    st.stop()
                # Read the converted file
                df = pl.read_excel(converted_file_path)
                # Clean up the temporary converted file
                try:
                    os.unlink(converted_file_path)
                except:
                    pass
            else:
                df = pl.read_excel(uploaded_file)
            new_columns = df.head(1).row(0)
            df = df.slice(0)
            df.columns = new_columns
            df = df.with_columns(
                pl.when(
                    df["Date"].str.starts_with("Aussie rules")
                )
                .then(df["Date"])
                .otherwise(None)
                .alias("League")
            ).with_columns(
                pl.col("League").forward_fill()
            )
            df = df.filter(pl.col("Postponed") == "0")
            filter_words = ["Australia.AFL"]
            df = df.filter(
                pl.col("League").str.contains("|".join(filter_words))
            )
            df = df.with_columns(
                pl.col("League")
                .str.replace(r",", "")
                .str.replace(r"(?i)\bweek\b", "")
                .str.replace("Aussie rules.", "")
                .str.strip_chars()
                .map_elements(
                    lambda x: re.sub(r'\d+', '', str(x)) if x is not None else None,
                    return_dtype=pl.Utf8
                )
                .alias("League")
            )
            # Filter out rows where League column contains "women" (case insensitive)
            if "League" in df.columns:
                if "League" in df.columns:
                    df = df.filter(
                        ~(
                            pl.col("League").str.to_lowercase().str.contains("Australia.SANFL") |
                            pl.col("League").str.contains("AFL Preseason")
                        )
    )
            columns_to_drop = ["1","2","3","4","OT","FT","Comment", "Postponed"]
            df = df.drop(columns_to_drop)
            if "Date" in df.columns:
                df = df.with_columns(
                    pl.col("Date").str.strptime(pl.Date, format="%d/%m %y")
                            .dt.strftime("%m/%d/%Y")
                            .alias("Date")
                )
            st.subheader("Processed League Data")
            df_display = df.select([
                "Date", "KO", "League", "Home", "Away", "Match Id"
            ])
            
            st.dataframe(df_display)
            current_date = datetime.now().strftime("%Y%m%d")
            
            # Download button - use df_display instead of df
            output = BytesIO()
            df_display.write_excel(output)  # Write the displayed columns only
            output.seek(0)
            st.download_button(
                label="Download Excel",
                data=output,
                file_name=f"Aussie Rules - {current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
        
    
elif page == "Program Review":

    def parse_sports_text(text: str) -> dict:
        """Parse sports text into structured data by extracting Sport, Category, Tournament, and Date.
        
        Args:
            text: Input string containing sports information with possible prefixes/suffixes
            
        Returns:
            Dictionary with keys: Sport, Category, Tournament, Date
            
        Raises:
            ValueError: If text doesn't contain at least Sport and Category separated by '-'
        """
        # Define patterns to remove (more maintainable as a list)
        patterns_to_remove = [
            "Assigned to Group - Competition creation: New season available -",
            "- /PROG. EXTENSION/ nan",
            "nan",
            "/PROG. EXTENSION/"
        ]
        
        # Clean the text by removing all specified patterns
        cleaned = text
        for pattern in patterns_to_remove:
            cleaned = cleaned.replace(pattern, "")
        cleaned = cleaned.strip()

        # Extract year (4-digit number in the tournament name)
        year_match = re.search(r'\b(20\d{2})\b', cleaned)
        year = int(year_match.group(1)) if year_match else 2025  # Default to 2025 if not found

        # Extract date (format like /2.6./)
        date_match = re.search(r'/(\d+)\.(\d+)\./', cleaned)
        date_str = ""
        if date_match:
            day, month = map(int, date_match.groups())
            try:
                date_obj = datetime(year=year, month=month, day=day)
                date_str = date_obj.strftime("%-m/%-d/%Y")  # Removes leading zeros (e.g., 6/2/2025)
            except ValueError:
                # Fallback if day/month is invalid (e.g., 31.2.)
                date_str = f"{month}/{day}/{year}"
        
        # Split into components, handling multiple hyphens and spaces
        parts = [part.strip() for part in cleaned.split("-") if part.strip()]
        
        # Validate we have at least Sport and Category
        if len(parts) < 2:
            raise ValueError(f"Invalid format. Expected 'Sport - Category [...]'. Got: '{text}'")
        
        # Extract components
        sport = parts[0]
        category = parts[1]
        tournament = " - ".join(parts[2:]) if len(parts) > 2 else ""
        
        # Additional cleaning of extracted values
        sport = sport.strip()
        category = category.strip()
        tournament = tournament.strip()
        
        # Validate required fields aren't empty after cleaning
        if not sport or not category:
            raise ValueError(f"Missing required fields. Sport: '{sport}', Category: '{category}'")
        
        return {
            "Sport": sport,
            "Category": category,
            "Tournament": tournament,  # Original Tournament text (unchanged)
            "Date": date_str  # Extracted date (e.g., "6/2/2025")
        }
    
    st.title("Sports Category Transformer")
    uploaded_file = st.file_uploader("Upload Excel file for Program Review", type=["xls", "xlsx", "csv"])
    
    if uploaded_file is not None:
        try:
            # Convert .xls to .xlsx if needed
            if uploaded_file.name.endswith('.xls'):
                st.info("Converting .xls file to .xlsx format...")
                converted_file_path = convert_xls_to_xlsx(uploaded_file)
                if converted_file_path is None:
                    st.error("Failed to convert .xls file. Please try again.")
                    st.stop()
                # Read the converted file
                df = pd.read_excel(converted_file_path, header=None)
                # Clean up the temporary converted file
                try:
                    os.unlink(converted_file_path)
                except:
                    pass
            else:
                # Read file based on type
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file, header=None)
                else:
                    df = pd.read_excel(uploaded_file, header=None)
            
            # Check if there's a second column and combine with first column
            if df.shape[1] > 1:
                df[0] = df[0].astype(str) + " " + df[1].astype(str)
                st.info("Combined data from multiple columns")
            
            # Process each row in the first column
            form_results = []
            for text in df.iloc[:, 0]:
                if pd.notna(text) and str(text).strip():
                    try:
                        result = parse_sports_text(str(text))
                        form_results.append(result)
                    except ValueError as e:
                        st.warning(f"Skipping row '{text[:50]}...': {e}")
            
            if form_results:
                df_display = pd.DataFrame(form_results)
                st.success(f"Successfully transformed {len(form_results)} rows!")
                st.dataframe(df_display)
                
                # Add download button
                current_date = datetime.now().strftime("%Y%m%d")
                csv = df_display.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Download as CSV",
                    data=csv,
                    file_name=f"program_review_{current_date}.csv",
                    mime='text/csv'
                )
            else:
                st.warning("No valid data found in the file.")
                
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
    else:
        st.info("Please upload a file to begin processing")
